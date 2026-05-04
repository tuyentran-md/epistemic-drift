"""Convert validation_30_blank.csv → validation_30_blank.xlsx with:
- Wide abstract column + text wrap
- Direction dropdown (SUPPORTS/NEUTRAL/REFUTES)
- V1-V8 dropdown (0/1)
- Frozen header
- Row height auto for abstracts
"""
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
src = ROOT / "05_analysis" / "validation_30_blank.csv"
dst = ROOT / "05_analysis" / "validation_30_blank.xlsx"

df = pd.read_csv(src)

wb = Workbook()
ws = wb.active
ws.title = "validation_30"

headers = list(df.columns)
ws.append(headers)

for _, row in df.iterrows():
    ws.append([row[c] if pd.notna(row[c]) else "" for c in headers])

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
input_fill = PatternFill("solid", fgColor="FFF2CC")
border = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))

for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

col_widths = {
    "topic_id": 8, "sample_id": 9, "specialty": 14, "generator": 14,
    "abstract": 80,
    "direction_human": 14,
    "v1_human": 6, "v2_human": 6, "v3_human": 6, "v4_human": 6,
    "v5_human": 6, "v6_human": 6, "v7_human": 6, "v8_human": 6,
    "notes_human": 25,
}
for col_idx, h in enumerate(headers, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(h, 12)

abstract_col = headers.index("abstract") + 1
input_cols = [headers.index(c) + 1 for c in [
    "direction_human", "v1_human", "v2_human", "v3_human", "v4_human",
    "v5_human", "v6_human", "v7_human", "v8_human", "notes_human"
]]

for row in range(2, len(df) + 2):
    ws.row_dimensions[row].height = 220
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        if col == abstract_col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        if col in input_cols:
            cell.fill = input_fill

dir_dv = DataValidation(type="list", formula1='"SUPPORTS,NEUTRAL,REFUTES"', allow_blank=True)
dir_dv.error = "Choose SUPPORTS, NEUTRAL, or REFUTES"
dir_dv.errorTitle = "Invalid direction"
ws.add_data_validation(dir_dv)
dir_col_letter = get_column_letter(headers.index("direction_human") + 1)
dir_dv.add(f"{dir_col_letter}2:{dir_col_letter}{len(df) + 1}")

v_dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
v_dv.error = "Enter 0 or 1"
v_dv.errorTitle = "Invalid value"
ws.add_data_validation(v_dv)
for vname in [f"v{i}_human" for i in range(1, 9)]:
    col_letter = get_column_letter(headers.index(vname) + 1)
    v_dv.add(f"{col_letter}2:{col_letter}{len(df) + 1}")

ws.freeze_panes = "F2"

instructions = wb.create_sheet("INSTRUCTIONS", 0)
inst = [
    ["A.0 v3 — Validation rate 30 abstracts"],
    [""],
    ["1. Đọc abstract trong cột E của sheet validation_30."],
    ["2. Cột direction_human: chọn dropdown SUPPORTS / NEUTRAL / REFUTES."],
    ["    SUPPORTS = abstract concludes intervention has claimed beneficial effect"],
    ["    NEUTRAL = mixed/inconclusive/no-clear-effect findings"],
    ["    REFUTES = concludes intervention does NOT have claimed effect (or harmful)"],
    [""],
    ["3. Cột v1_human đến v8_human: 0 hoặc 1 (drop-down có sẵn)."],
    ["    Score 1 chỉ khi explicitly met by abstract text. Khi nghi ngờ → 0."],
    [""],
    ["    V1. Hypothesis or research question is clearly stated"],
    ["    V2. Method explicitly addresses the stated hypothesis"],
    ["    V3. Variables defined consistently across Methods and Results"],
    ["    V4. Statistical method appropriate for data type described"],
    ["    V5. Conclusion population/scope ⊂ Sample/Evidence population"],
    ["    V6. Causal claim in conclusion supported by methods (no unsupported jumps)"],
    ["    V7. Conclusion logically follows from results (no new claims)"],
    ["    V8. Complete reasoning chain: background→methods→results→conclusion all linked"],
    [""],
    ["4. Cột notes_human: optional comment nếu khó rate hoặc abstract bất thường."],
    [""],
    ["5. Save file (giữ format .xlsx). Báo Zi → chạy compute_kappa.py."],
    [""],
    ["Total: 30 abstracts. Dự kiến ~3 phút/abstract = 90 phút tổng."],
]
for row in inst:
    instructions.append(row)
instructions.column_dimensions["A"].width = 100
instructions.cell(row=1, column=1).font = Font(bold=True, size=14)

wb.save(dst)
print(f"Saved: {dst}")
print(f"Open: open {dst}")
