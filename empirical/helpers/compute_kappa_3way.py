"""Compute pairwise Cohen's κ across 3 raters: sếp, Zi, AI (DeepSeek + Claude).

Per OSF X4RP5 §4 H5: human (sếp) vs LLM rater κ ≥ 0.7 substantial agreement.
Adds Zi as a second human-style rater for cross-validation.

Run after sếp + Zi rated 30 abstracts:
    python helpers/compute_kappa_3way.py
"""
import json
from pathlib import Path
import pandas as pd
from sklearn.metrics import cohen_kappa_score, accuracy_score
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
ANA = ROOT / "05_analysis"


def load_sep():
    wb = openpyxl.load_workbook(ANA / "validation_30_blank.xlsx", data_only=True)
    ws = wb["validation_30"]
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True)]
    cols = [c.value for c in ws[1]]
    df = pd.DataFrame(rows, columns=cols)
    df["key"] = df["topic_id"] + "_" + df["sample_id"].astype(int).astype(str) + "_" + df["generator"].astype(str)
    return df


def load_zi():
    df = pd.read_csv(ANA / "validation_30_zi.csv")
    df["key"] = df["topic_id"] + "_" + df["sample_id"].astype(int).astype(str) + "_" + df["generator"].astype(str)
    return df


def load_ai_for_keys(keys):
    """Load AI ratings matched by key. Generator label normalised."""
    gen_map = {"gemini-3-pro": "gemini-3-pro-preview",
               "gemma3-12b": "gemma3_12b",
               "llama3.2-3b": "llama3.2_3b"}
    ai_rows = {}
    for label, slug in gen_map.items():
        p = ANA / f"rater_ai__{slug}.jsonl"
        with open(p) as fh:
            for line in fh:
                r = json.loads(line)
                key = f"{r['topic_id']}_{r['sample_id']}_{label}"
                ai_rows[key] = r
    out = []
    for k in keys:
        r = ai_rows.get(k)
        if r is None:
            continue
        out.append({
            "key": k,
            "direction_ai": r.get("direction"),
            "rater_ai": r.get("rater_model"),
            **{f"v{i}_ai": r.get(f"v{i}") for i in range(1, 9)},
        })
    return pd.DataFrame(out)


def landis_koch(k):
    if k is None or pd.isna(k):
        return "n/a"
    if k < 0:
        return "poor (<0)"
    if k <= 0.20:
        return "slight (0–0.20)"
    if k <= 0.40:
        return "fair (0.21–0.40)"
    if k <= 0.60:
        return "moderate (0.41–0.60)"
    if k <= 0.80:
        return "substantial (0.61–0.80)"
    return "almost perfect (0.81–1.00)"


def kappa_or_none(a, b):
    if len(a) < 5:
        return None
    try:
        return cohen_kappa_score(a, b)
    except Exception:
        return None


def main():
    sep = load_sep()
    zi = load_zi()
    ai = load_ai_for_keys(sep["key"].tolist())
    merged = sep.merge(zi[["key","direction_zi"] + [f"v{i}_zi" for i in range(1,9)]], on="key")
    merged = merged.merge(ai, on="key")
    print(f"Merged rows: {len(merged)}")
    # Drop rows where AI rater failed (no direction)
    before = len(merged)
    merged = merged[merged["direction_ai"].notna()].copy()
    if len(merged) < before:
        print(f"  Dropped {before - len(merged)} rows with missing AI direction (rater errors).")
    # Coerce all V cols to numeric, drop unrecoverable
    for i in range(1, 9):
        for col in [f"v{i}_human", f"v{i}_zi", f"v{i}_ai"]:
            merged[col] = pd.to_numeric(merged[col], errors="coerce")

    # Direction labels — uppercase strip
    sep_d = merged["direction_human"].astype(str).str.strip().str.upper()
    zi_d = merged["direction_zi"].astype(str).str.strip().str.upper()
    ai_d = merged["direction_ai"].astype(str).str.strip().str.upper()

    labels = ["SUPPORTS", "NEUTRAL", "REFUTES"]
    md = ["# Three-rater agreement (sếp, Zi, AI rater)",
          f"_Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}_",
          f"_n abstracts: {len(merged)}_", ""]

    md.append("## Direction (3-class)")
    pairs = [
        ("Sếp vs Zi (human-human)", sep_d, zi_d),
        ("Sếp vs AI", sep_d, ai_d),
        ("Zi vs AI", zi_d, ai_d),
    ]
    for name, a, b in pairs:
        k = cohen_kappa_score(a, b, labels=labels)
        acc = accuracy_score(a, b)
        md.append(f"- **{name}:** κ = {k:.3f} ({landis_koch(k)}), accuracy = {acc:.1%}")
    md.append("")

    # Per-rater AI breakdown
    md.append("## Direction κ by AI rater subset")
    for ai_model in merged["rater_ai"].dropna().unique():
        sub = merged[merged["rater_ai"] == ai_model]
        if len(sub) < 5:
            md.append(f"- {ai_model} (n={len(sub)}): too few")
            continue
        sub_sep = sub["direction_human"].astype(str).str.strip().str.upper()
        sub_zi = sub["direction_zi"].astype(str).str.strip().str.upper()
        sub_ai = sub["direction_ai"].astype(str).str.strip().str.upper()
        k_sa = cohen_kappa_score(sub_sep, sub_ai, labels=labels)
        k_za = cohen_kappa_score(sub_zi, sub_ai, labels=labels)
        md.append(f"- **{ai_model}** (n={len(sub)}): sếp-AI κ = {k_sa:.3f}, Zi-AI κ = {k_za:.3f}")
    md.append("")

    # V1-V8 binary kappa
    md.append("## V1-V8 binary item agreement")
    md.append("| Item | sếp-Zi | sếp-AI | Zi-AI |")
    md.append("|---|---|---|---|")
    sep_v_means = []
    zi_v_means = []
    ai_v_means = []
    for i in range(1, 9):
        col_h = f"v{i}_human"
        col_z = f"v{i}_zi"
        col_a = f"v{i}_ai"
        sub = merged.dropna(subset=[col_h, col_z, col_a]).copy()
        if len(sub) < 5:
            md.append(f"| V{i} | n<5 | n<5 | n<5 |")
            continue
        h = sub[col_h].astype(float).astype(int).tolist()
        z = sub[col_z].astype(float).astype(int).tolist()
        a = sub[col_a].astype(float).astype(int).tolist()
        k_sz = kappa_or_none(h, z)
        k_sa = kappa_or_none(h, a)
        k_za = kappa_or_none(z, a)
        sep_v_means.append(sum(h) / len(h))
        zi_v_means.append(sum(z) / len(z))
        ai_v_means.append(sum(a) / len(a))
        f = lambda k: f"{k:.3f}" if k is not None else "—"
        md.append(f"| V{i} | {f(k_sz)} | {f(k_sa)} | {f(k_za)} |")
    md.append("")
    md.append(f"Mean pass rate per V-item:")
    md.append(f"- sếp: {[f'{v:.2f}' for v in sep_v_means]}")
    md.append(f"- Zi: {[f'{v:.2f}' for v in zi_v_means]}")
    md.append(f"- AI: {[f'{v:.2f}' for v in ai_v_means]}")
    md.append("")

    # V-score continuous
    sep_vsum = merged[[f"v{i}_human" for i in range(1,9)]].astype(float).sum(axis=1)
    zi_vsum = merged[[f"v{i}_zi" for i in range(1,9)]].astype(int).sum(axis=1)
    ai_vsum = merged[[f"v{i}_ai" for i in range(1,9)]].astype(int).sum(axis=1)
    md.append("## V-score sum (0-8) per rater")
    md.append(f"- sếp mean = {sep_vsum.mean():.2f}, sd = {sep_vsum.std():.2f}")
    md.append(f"- Zi mean = {zi_vsum.mean():.2f}, sd = {zi_vsum.std():.2f}")
    md.append(f"- AI mean = {ai_vsum.mean():.2f}, sd = {ai_vsum.std():.2f}")
    from scipy import stats
    md.append(f"- Pearson r (sếp, Zi) = {stats.pearsonr(sep_vsum, zi_vsum)[0]:.3f}")
    md.append(f"- Pearson r (sếp, AI) = {stats.pearsonr(sep_vsum, ai_vsum)[0]:.3f}")
    md.append(f"- Pearson r (Zi, AI)  = {stats.pearsonr(zi_vsum, ai_vsum)[0]:.3f}")
    md.append("")

    # Confusion matrix sep vs ai (key H5)
    md.append("## Confusion matrix — Sếp vs AI (rows=sếp, cols=AI)")
    cm = pd.crosstab(sep_d, ai_d, margins=True)
    md.append("```")
    md.append(cm.to_string())
    md.append("```\n")

    # Decision per H5
    sep_ai_k = cohen_kappa_score(sep_d, ai_d, labels=labels)
    zi_ai_k = cohen_kappa_score(zi_d, ai_d, labels=labels)
    md.append("## Decision per OSF X4RP5 §4 H5")
    md.append(f"- Threshold: human-AI κ ≥ 0.7 (substantial)")
    md.append(f"- **Sếp-AI κ = {sep_ai_k:.3f}** → {'PASS ✅' if sep_ai_k >= 0.7 else 'BELOW threshold ❌'}")
    md.append(f"- Zi-AI κ = {zi_ai_k:.3f} (cross-check)")
    md.append("")
    md.append("Notes: AI rater is mixed DeepSeek V3.1 + Claude Sonnet 4.5 (50/50 deterministic split). "
              "Per-rater κ shown above. Per OSF prereg, both LLM rater subsets must reach κ ≥ 0.7 "
              "for full H5 confirmation.")

    out = ANA / "validation_kappa_3way.md"
    out.write_text("\n".join(md))
    print(f"\nReport: {out}")
    print(f"\nKey result: Sếp-AI κ = {sep_ai_k:.3f} ({landis_koch(sep_ai_k)})")
    print(f"            Zi-AI κ = {zi_ai_k:.3f}")
    print(f"            Sếp-Zi κ = {cohen_kappa_score(sep_d, zi_d, labels=labels):.3f}")


if __name__ == "__main__":
    main()
